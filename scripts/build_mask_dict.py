"""キュレーション済み Excel から マスク辞書 YAML（data/mask_dict.yaml 形式）を生成する
使い捨てスクリプト。登録時に **冗長エントリのチェック** も掛ける。

着想:
  用語集や find_feature_names / classify_glossary_terms の候補を人手でレビューして
  Excel に落とし込んだあと、それを手作業で YAML に書き写すのは事故のもと。
  「用語・分類・部分一致 だけ埋めた Excel」→「そのまま読める辞書 YAML」に機械変換する。

前提フォーマット（Excel。列は**列位置**で指定。既定 A/D/E/F）:
    A 列: 用語        … 登録する表層（canonical になる）
    D 列: 種類        … Company / Trademark / Person（→ 社名 / 商標 / 人名）
    E 列: 部分一致    … 真偽値（○/TRUE/1/✓/はい 等が入っていれば 部分一致: true）
    F 列: 大小区別    … 真偽値（略語向け。true なら STS は STS のみ・Sts/sts は拾わない）
  1 行目はヘッダとして読み飛ばす（--header-row で変更可）。

「部分一致」とは（src/masking/dictionary.py と同義）:
  - **既定（空欄）＝まるごと一致**: その用語がまるごとの語として現れたときだけ隠す。
  - **部分一致: true**: 他の語の**中**に境界沿いで現れても、その部分を隠す。
    区切り入り（IF- → IF-X の「IF-」）でも camelCase（iAS → iASMap の「iAS」）でも、
    同じ 部分一致: true で効く（embed/接頭辞という実装区別は不要）。

冗長性チェック（登録時）:
  - **完全重複**: 正規化キーが既出（iAS が2回、iAS と全角 ｉＡＳ 等）→ **自動除去**。
  - **部分一致が過剰に広い**: 部分一致: true なのにキーが1文字（例 A）→ **警告のみ**
    （境界沿いとはいえ広く巻き込むため。登録は残す＝recall 優先。人手で見直す用）。

  ※ 正規化規則は src/masking/dictionary.py の normalize と論理一致させてある（self-contained の
     ため写経）。dictionary.py 側を変えたらここも合わせること。

使い方:
    # dry-run（既定。チェックとプレビューのみ・書き出さない）
    uv run python scripts/build_mask_dict.py <辞書.xlsx>

    # 別ファイルへ書き出す（完全重複を自動除去して警告）
    uv run python scripts/build_mask_dict.py <辞書.xlsx> --out data/mask_dict.generated.yaml

    # 列位置・ヘッダ行・シートを変える
    uv run python scripts/build_mask_dict.py <辞書.xlsx> --term-col A --kind-col D \
        --partial-col E --case-col F --header-row 1 --sheet Sheet1
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import yaml

# --------------------------------------------------------------------------- #
# 正規化（src/masking/dictionary.py の normalize と論理一致させた写し）
# --------------------------------------------------------------------------- #

_WHITESPACE = re.compile(r"\s+")


def normalize(text: str) -> str:
    """照合用の正規化（NFKC で全角半角統一・casefold で大小無視・空白除去）。"""
    return _WHITESPACE.sub("", unicodedata.normalize("NFKC", text).casefold())


# --------------------------------------------------------------------------- #
# Excel 入力
# --------------------------------------------------------------------------- #

# 種類（D 列）→ 辞書セクション。英日どちらの表記も受ける。casefold 済みで引く。
_KIND_TO_SECTION = {
    "company": "社名",
    "trademark": "商標",
    "person": "人名",
    "社名": "社名",
    "会社名": "社名",
    "商標": "商標",
    "人名": "人名",
    "社員名": "人名",
}
_FALLBACK_SECTION = "商標"  # 種類が空/不明のときの退避先（捨てない＝recall 優先）

# 部分一致（E 列）の真偽解釈。casefold・NFKC・strip 済みで突き合わせる。
_TRUE_VALUES = {
    "1",
    "true",
    "t",
    "yes",
    "y",
    "○",
    "◯",
    "〇",
    "✓",
    "✔",
    "はい",
    "有",
    "x",
}
_FALSE_VALUES = {"", "0", "false", "f", "no", "n", "×", "-", "なし", "無"}


@dataclass
class RawRow:
    """Excel 1 行分（生値）。"""

    term: str
    kind: str | None
    partial_raw: str | None
    case_raw: str | None
    row: int


def _col_to_index(letter: str) -> int:
    """列文字（A/D/E …）を 0 始まりインデックスへ。"""
    from openpyxl.utils import column_index_from_string  # type: ignore[import-untyped]

    return column_index_from_string(letter.strip().upper()) - 1


def _cell(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def load_rows(
    path: Path,
    sheet: str | None,
    header_row: int,
    term_col: str,
    kind_col: str,
    partial_col: str,
    case_col: str,
) -> list[RawRow]:
    """Excel を読み、用語のある行を RawRow として返す。"""
    from openpyxl import load_workbook  # type: ignore[import-untyped]  # スタブ無し

    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet is not None:
        if sheet not in wb.sheetnames:
            raise SystemExit(
                f"エラー: シート {sheet!r} が見つかりません（存在: {wb.sheetnames}）"
            )
        ws = wb[sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    ti, ki, pi, ci = (
        _col_to_index(term_col),
        _col_to_index(kind_col),
        _col_to_index(partial_col),
        _col_to_index(case_col),
    )

    def pick(row: tuple, idx: int) -> str | None:
        return _cell(row[idx]) if len(row) > idx else None

    rows: list[RawRow] = []
    for n, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        term = pick(row, ti)
        if term is None:
            continue
        rows.append(
            RawRow(
                term=term,
                kind=pick(row, ki),
                partial_raw=pick(row, pi),
                case_raw=pick(row, ci),
                row=n,
            )
        )
    return rows


# --------------------------------------------------------------------------- #
# 行 → 辞書エントリ
# --------------------------------------------------------------------------- #


@dataclass
class Entry:
    """辞書 1 エントリ（このスクリプトが扱う最小形。aliases/mask は持たない）。"""

    canonical: str
    section: str
    partial: bool
    case_sensitive: bool
    row: int


def parse_bool(raw: str | None) -> tuple[bool, bool]:
    """真偽列の解釈。返り値 (value, unknown)。unknown=True なら未知値（False 扱い）。"""
    if raw is None:
        return False, False
    key = normalize(raw)
    if key in _TRUE_VALUES:
        return True, False
    if key in _FALSE_VALUES:
        return False, False
    return False, True  # 未知値は安全側（off）＋警告


def to_entries(rows: list[RawRow]) -> tuple[list[Entry], list[str]]:
    """RawRow 群を Entry 群へ。種類→セクション変換・部分一致/大小区別の解釈をし、警告を集める。"""
    entries: list[Entry] = []
    warnings: list[str] = []
    for r in rows:
        kind_key = normalize(r.kind) if r.kind else ""
        section = _KIND_TO_SECTION.get(kind_key)
        if section is None:
            section = _FALLBACK_SECTION
            warnings.append(
                f"行{r.row}: 種類 {r.kind!r} を解釈できず {_FALLBACK_SECTION} に退避 "
                f"（{r.term!r}）。Company/Trademark/Person を入れてください。"
            )
        partial, p_unknown = parse_bool(r.partial_raw)
        if p_unknown:
            warnings.append(
                f"行{r.row}: 部分一致 値 {r.partial_raw!r} を解釈できず false 扱い（{r.term!r}）。"
            )
        case_sensitive, c_unknown = parse_bool(r.case_raw)
        if c_unknown:
            warnings.append(
                f"行{r.row}: 大小区別 値 {r.case_raw!r} を解釈できず false 扱い（{r.term!r}）。"
            )
        entries.append(
            Entry(
                canonical=r.term,
                section=section,
                partial=partial,
                case_sensitive=case_sensitive,
                row=r.row,
            )
        )
    return entries, warnings


# --------------------------------------------------------------------------- #
# 冗長性チェック
# --------------------------------------------------------------------------- #


@dataclass
class Finding:
    """冗長性の指摘 1 件。"""

    row: int
    canonical: str
    kind: str  # DUP / PARTIAL_BROAD
    message: str
    dropped: bool


@dataclass
class CheckResult:
    kept: list[Entry] = field(default_factory=list)
    findings: list[Finding] = field(default_factory=list)


def check_redundancies(entries: list[Entry]) -> CheckResult:
    """登録エントリの冗長性を検査し、自動処理後の kept と findings を返す。

    - **完全重複**（正規化キーが既出）→ 除去。
    - **部分一致が過剰に広い**（部分一致 かつ キーが1文字）→ 警告のみ（登録は残す）。
    """
    result = CheckResult()
    seen: set[str] = set()

    for e in entries:
        key = normalize(e.canonical)
        if key in seen:
            result.findings.append(
                Finding(
                    e.row,
                    e.canonical,
                    "DUP",
                    "正規化キーが既出（大小/全角半角/空白の違いを吸収すると同一）→ 除去",
                    dropped=True,
                )
            )
            continue
        if e.partial and len(key) <= 1:
            result.findings.append(
                Finding(
                    e.row,
                    e.canonical,
                    "PARTIAL_BROAD",
                    "部分一致なのにキーが1文字＝広く巻き込む恐れ。意図を確認（登録は残す）",
                    dropped=False,
                )
            )
        seen.add(key)
        result.kept.append(e)

    return result


# --------------------------------------------------------------------------- #
# YAML 出力（dictionary.save_entries と同じ**フルセット英語**書式を self-contained に再現）
# --------------------------------------------------------------------------- #

# 内部カテゴリ（社名/商標/人名）→ 英語セクション名。dictionary._CATEGORY_SECTION と一致させる。
_CATEGORY_SECTION = {"社名": "Company", "商標": "Trademark", "人名": "Person"}
_SECTION_ORDER = ("Company", "Trademark", "Person")


def build_yaml_tree(entries: list[Entry]) -> dict[str, list]:
    """Entry 群を {英語セクション: [フルセット dict]} に畳む（各節を正規化キーでソート）。

    dictionary.save_entries と同じ書式：各エントリは canonical/aliases/mask/partial/case_sensitive
    の全キーを持ち、未定義は null（真偽値は false）。このスクリプトは aliases/mask を持たないので常に null。
    """
    by_section: dict[str, list[tuple[str, dict]]] = defaultdict(list)
    for e in entries:
        section = _CATEGORY_SECTION.get(e.section, e.section)
        obj = {
            "canonical": e.canonical,
            "aliases": None,
            "mask": None,
            "partial": e.partial,
            "case_sensitive": e.case_sensitive,
        }
        by_section[section].append((normalize(e.canonical), obj))

    sorted_sections = {
        s: [item for _, item in sorted(rows, key=lambda r: r[0])]
        for s, rows in by_section.items()
    }
    ordered = {s: sorted_sections[s] for s in _SECTION_ORDER if s in sorted_sections}
    ordered.update({s: v for s, v in sorted_sections.items() if s not in ordered})
    return ordered


def dump_yaml(entries: list[Entry]) -> str:
    return yaml.safe_dump(
        build_yaml_tree(entries), allow_unicode=True, sort_keys=False, indent=2
    )


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

_KIND_LABEL = {
    "DUP": "完全重複",
    "PARTIAL_BROAD": "部分一致が広い",
}
_LIST_CAP = 10  # 各カテゴリの明細表示の上限（超えたら「…他 N 件」）


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")  # Windows コンソール文字化け対策

    parser = argparse.ArgumentParser(
        description="キュレーション済み Excel から マスク辞書 YAML を生成し、冗長登録を検査する。"
    )
    parser.add_argument("excel", type=Path, help="入力 Excel(.xlsx/.xlsm)")
    parser.add_argument("--sheet", default=None, help="シート名（既定: 先頭シート）")
    parser.add_argument(
        "--header-row", type=int, default=1, help="ヘッダ行(1-based。既定 1)"
    )
    parser.add_argument("--term-col", default="A", help="用語の列（既定 A）")
    parser.add_argument("--kind-col", default="D", help="種類の列（既定 D）")
    parser.add_argument("--partial-col", default="E", help="部分一致の列（既定 E）")
    parser.add_argument("--case-col", default="F", help="大小区別の列（既定 F）")
    parser.add_argument(
        "--out",
        type=Path,
        help="書き出し先 YAML（未指定なら dry-run＝チェックとプレビューのみ）",
    )
    args = parser.parse_args(argv)

    if not args.excel.exists():
        raise SystemExit(f"エラー: Excel が存在しません: {args.excel}")

    rows = load_rows(
        args.excel,
        args.sheet,
        args.header_row,
        args.term_col,
        args.kind_col,
        args.partial_col,
        args.case_col,
    )
    print(f"入力: {args.excel}（用語 {len(rows)} 件）")

    entries, warnings = to_entries(rows)
    for w in warnings:
        print(f"  ⚠ {w}")

    result = check_redundancies(entries)

    by_kind: dict[str, list[Finding]] = defaultdict(list)
    for f in result.findings:
        by_kind[f.kind].append(f)
    dropped = by_kind.get("DUP", [])
    broad = by_kind.get("PARTIAL_BROAD", [])

    print("\n== 冗長性チェック ==")
    if not result.findings:
        print("  指摘なし。")

    if dropped:
        print(f"\n■ 完全重複 {len(dropped)} 件（自動で除去）")
        for f in dropped[:_LIST_CAP]:
            print(f"    ・行{f.row} {f.canonical!r}")
        if len(dropped) > _LIST_CAP:
            print(f"    … 他 {len(dropped) - _LIST_CAP} 件")

    if broad:
        print(f"\n■ 部分一致が広い {len(broad)} 件（要確認・登録は残す）")
        for f in broad[:_LIST_CAP]:
            print(f"    ・行{f.row} {f.canonical!r} — {f.message}")
        if len(broad) > _LIST_CAP:
            print(f"    … 他 {len(broad) - _LIST_CAP} 件")

    # セクション別件数（除去後）。counts は内部カテゴリ（社名/商標/人名）でキーする。
    counts: dict[str, int] = defaultdict(int)
    partial_counts: dict[str, int] = defaultdict(int)
    cs_counts: dict[str, int] = defaultdict(int)
    for e in result.kept:
        counts[e.section] += 1
        if e.partial:
            partial_counts[e.section] += 1
        if e.case_sensitive:
            cs_counts[e.section] += 1
    _cat_order = ("社名", "商標", "人名")
    summary = " / ".join(
        f"{s} {counts[s]}件(部分一致 {partial_counts[s]}/大小区別 {cs_counts[s]})"
        for s in _cat_order
        if counts[s]
    )
    other = " / ".join(f"{s} {counts[s]}件" for s in counts if s not in _cat_order)
    print(
        f"\n登録（除去後 {len(result.kept)} 件）: {summary}{(' / ' + other) if other else ''}"
    )

    yaml_text = dump_yaml(result.kept)

    if args.out is None:
        print("\n----- プレビュー（dry-run。書き出すには --out を付ける）-----")
        print(yaml_text.rstrip())
    else:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(yaml_text, encoding="utf-8")
        print(f"\n書き出しました: {args.out.resolve()}")
        print("  ※ 本体 data/mask_dict.yaml へは人手で確認のうえマージしてください。")

    return 0


if __name__ == "__main__":
    sys.exit(main())
