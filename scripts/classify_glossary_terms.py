"""用語集(Excel)の各用語について「概要・メモ」等を LLM に読ませ、
その用語が **秘匿すべき（マスク対象）か** を根拠つきで判定する使い捨てスクリプト。

着想:
  用語単体（表層）だけでは「自社固有の機能名か／業界一般語か」は判断できない
  （社内ジャーゴンは関根の拡張固有表現体系に無く NER 単体では引けない）。
  だが用語集の「概要・メモ」列には、その用語が何を指すかの説明がある。
  ここを LLM に読ませれば「開示すると自社の技術・製品情報が漏れる固有名か」を
  文脈から判断できる、という仮説。find_feature_names.py（ファイル名ヒット＝使用実績）
  とは相補的で、両者を突き合わせると確度が上がる。

前提フォーマット（実ファイル基準）:
  シート「用語集」・6 行目ヘッダ・列
    C: 用語 / D: 英語名称 / E: 日本語名称 / F: 種類 / G: 概要・メモ
  （列は**ヘッダ名で探す**ので、位置がずれても・欠けても動く。用語列さえあれば可。）

判定:
  各用語を Azure OpenAI gpt-4.1-mini に投げ、以下を返させる。
    - mask   : 要マスク / 不要 / 要確認（判断材料不足）
    - confidence : 高 / 中 / 低（＝「秘匿すべき」とどれだけ確実に言えるか）
    - category   : 製品名/機能名/コードネーム/型番/社内固有 などの分類（LLM の自由記述）
    - reason     : 根拠（なぜそう判定したか。主に概要・メモに基づく）
  precision 優先: **確実に秘匿すべきものを高純度で選び出す** のが狙い。曖昧なものは
  要マスクに水増しせず、固有だと明確に言い切れるものだけを要マスクにする。
  「要マスク × 確信度高」＝**確実に秘匿すべき確定候補**として最上段に出す。
  ※これは候補であって確定辞書ではない。最終判断は必ず人手レビュー。

送信先の制約:
  用語集そのものが機密（自社固有名の一覧）になり得るため、送信先は PII と同じく
  **Azure OpenAI gpt-4.1-mini（承認済みエンドポイント）に限定**する。汎用 API に投げない。
  実体は pii-masker の client.get_client を再利用（DefaultAzureCredential / az login）。

使い方:
  # 実機（Azure が使える別マシン。要 az login ＋ .env の RESOURCE_NAME_GPT41_MINI）
  uv run python scripts/classify_glossary_terms.py <用語集.xlsm> --out out.csv

  # まず少数で試す（コスト確認）
  uv run python scripts/classify_glossary_terms.py <用語集.xlsm> --limit 30 --out try.csv

  # LLM 無しで配線だけ検証（このマシン用。判定は簡易ヒューリスティックの偽物）
  uv run python scripts/classify_glossary_terms.py <用語集.xlsm> --mock --out out.csv
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
from collections.abc import Callable, Sequence
from dataclasses import dataclass
from pathlib import Path

# --------------------------------------------------------------------------- #
# 用語集(Excel)パース（find_feature_names.py と同じ探索方針。列を 2 つ追加）
# --------------------------------------------------------------------------- #

GLOSSARY_SHEET_NAME = "用語集"
GLOSSARY_HEADER_ROW = 6
COL_TERM = "用語"
COL_EN = "英語名称"
COL_JA = "日本語名称"
COL_KIND = "種類"
COL_MEMO = "概要・メモ"
_ALL_COLS = (COL_TERM, COL_EN, COL_JA, COL_KIND, COL_MEMO)
_HEADER_SEARCH_LIMIT = 30  # 指定行にヘッダが無いときに走査する行数


@dataclass
class GlossaryRow:
    """判定対象 1 行分。用語列以外は欠けても None のまま続行する。"""

    term: str  # 「用語」列の原文（照合キーではなく、そのまま表示・LLM へ渡す）
    english: str | None
    japanese: str | None
    kind: str | None  # 種類
    memo: str | None  # 概要・メモ（判定の主材料）
    row: int


def _cell(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _find_header(sheet, header_row: int) -> tuple[int, dict[str, int]] | None:
    """ヘッダ行番号と列インデックスの対応を返す。まず指定行、無ければ先頭数十行を探索。"""
    candidates = [header_row] + [
        r for r in range(1, _HEADER_SEARCH_LIMIT + 1) if r != header_row
    ]
    for r in candidates:
        row = next(sheet.iter_rows(min_row=r, max_row=r, values_only=True), None)
        if row is None:
            continue
        header_map: dict[str, int] = {}
        for idx, cell in enumerate(row):
            name = _cell(cell)
            if name in _ALL_COLS and name not in header_map:
                header_map[name] = idx
        if COL_TERM in header_map:
            return r, header_map
    return None


def load_glossary_rows(
    path: Path, sheet_name: str, header_row: int
) -> tuple[list[GlossaryRow], list[str]]:
    """用語集 Excel から GlossaryRow 群を読む。「用語」列さえあれば読める。"""
    from openpyxl import load_workbook  # type: ignore[import-untyped]  # スタブ無し

    warnings: list[str] = []
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"エラー: シート {sheet_name!r} が見つかりません（存在: {wb.sheetnames}）"
        )
    sheet = wb[sheet_name]

    found = _find_header(sheet, header_row)
    if found is None:
        raise SystemExit(
            f"エラー: ヘッダ行に {COL_TERM!r} 列が見つかりません"
            f"（{header_row} 行目と先頭 {_HEADER_SEARCH_LIMIT} 行を確認）"
        )
    actual_row, header_map = found
    if actual_row != header_row:
        warnings.append(f"ヘッダを {header_row} 行目でなく {actual_row} 行目で検出。")
    for col in (COL_EN, COL_JA, COL_KIND, COL_MEMO):
        if col not in header_map:
            warnings.append(f"{col!r} 列なし（その情報なしで続行）。")
    if COL_MEMO not in header_map:
        warnings.append(
            "⚠ 概要・メモ列が無い＝判定の主材料が欠ける。用語名だけの弱い判定になる。"
        )

    ti = header_map[COL_TERM]
    ei = header_map.get(COL_EN)
    ji = header_map.get(COL_JA)
    ki = header_map.get(COL_KIND)
    mi = header_map.get(COL_MEMO)

    def pick(row: tuple, idx: int | None) -> str | None:
        return _cell(row[idx]) if idx is not None and len(row) > idx else None

    rows: list[GlossaryRow] = []
    for n, row in enumerate(
        sheet.iter_rows(min_row=actual_row + 1, values_only=True), start=actual_row + 1
    ):
        term = _cell(row[ti]) if len(row) > ti else None
        if term is None:
            continue
        rows.append(
            GlossaryRow(
                term=term,
                english=pick(row, ei),
                japanese=pick(row, ji),
                kind=pick(row, ki),
                memo=pick(row, mi),
                row=n,
            )
        )
    return rows, warnings


# --------------------------------------------------------------------------- #
# LLM 判定
# --------------------------------------------------------------------------- #

DEFAULT_MODEL = "gpt-4.1-mini"
DEFAULT_BATCH_SIZE = 25

# 判定値の正規化テーブル（LLM の表記ゆれを吸収）。
_MASK_CANON = {
    "yes": "要マスク",
    "mask": "要マスク",
    "要マスク": "要マスク",
    "no": "不要",
    "keep": "不要",
    "不要": "不要",
    "unsure": "要確認",
    "review": "要確認",
    "要確認": "要確認",
}
_CONF_CANON = {
    "high": "高",
    "高": "高",
    "mid": "中",
    "medium": "中",
    "中": "中",
    "low": "低",
    "低": "低",
}

_PROMPT_HEADER = """\
あなたは、ある製造業・技術系企業の社内文書のマスキング（秘匿）を支援する専門家です。
以下は自社の「用語集」の抜粋です。目的は、**確実に秘匿すべき（＝外部開示で自社の製品・技術・
組織の情報が漏れる）自社固有の語を、高い精度で選び出す**ことです。曖昧なものを要マスクに
水増ししないでください（要マスク集合は「確実なものの名簿」であってほしい）。

判定材料は主に「概要・メモ」（その用語の説明）です。種類・英語名称・日本語名称も参考にします。

判定（mask）:
- 要マスク(yes): 概要・メモから、**自社固有の**製品名・機能名・コードネーム・型番・
  社内プロジェクト名・独自アルゴリズム/方式の名称だと **明確に読み取れる** もの。
- 不要(no): 業界・学術で一般に通用する技術用語、汎用的な一般名詞、標準規格名など、
  自社に固有でなく開示しても情報漏洩にならない語。
- 要確認(unsure): 概要・メモが乏しい等で、固有か一般か **決められない** 語。

確信度（confidence）＝「秘匿すべき」とどれだけ確実に言えるか:
- high: 概要・メモに固有性の **明確な根拠** がある（「自社独自」「専用」「社内」「コードネーム」
  「当社の」「プロジェクト◯◯」等、または明らかに一社固有の製品/機能名）。＝確実に秘匿すべき。
- mid: 固有の可能性が高いが、根拠がやや弱い（要レビュー）。
- low: 手がかりが薄い。

最重要の原則（precision 優先）:
- **迷ったら要マスクに入れない**。固有だと **明確に言い切れるものだけ** を要マスクにする。
- 特に "要マスク × high" は「確実に秘匿すべき」の確定候補なので、根拠が曖昧なら high にしない。
- reason には、そう判定した **具体的な根拠を概要・メモから引いて** 書く
  （high のときは「何が固有性の証拠か」を必ず明記する）。

出力は JSON オブジェクト 1 個のみ。キー "results" に配列を入れる。
各要素は {"no": 用語の番号(整数), "mask": "yes"|"no"|"unsure",
"confidence": "high"|"mid"|"low", "category": "分類の短い語", "reason": "判定の根拠(日本語1〜2文)"}。
入力の全番号について、過不足なく 1 要素ずつ返すこと。

# 用語一覧
"""


def build_prompt(batch: Sequence[GlossaryRow]) -> str:
    """番号つきの用語一覧を組み、判定指示に連結した完成プロンプトを返す。"""
    lines: list[str] = []
    for i, r in enumerate(batch, start=1):
        parts = [f"[{i}] 用語: {r.term}"]
        if r.english:
            parts.append(f"英語名称: {r.english}")
        if r.japanese:
            parts.append(f"日本語名称: {r.japanese}")
        if r.kind:
            parts.append(f"種類: {r.kind}")
        parts.append(f"概要・メモ: {r.memo or '(記載なし)'}")
        lines.append("\n  ".join(parts))
    return _PROMPT_HEADER + "\n\n".join(lines)


@dataclass
class Judgement:
    mask: str  # 要マスク / 不要 / 要確認
    confidence: str  # 高 / 中 / 低
    category: str
    reason: str


def _unsure(reason: str) -> Judgement:
    """判断不能時のフェイルセーフ（recall 優先で要確認へ寄せる）。"""
    return Judgement("要確認", "低", "", reason)


def _parse_results(raw: str, n: int) -> list[Judgement]:
    """LLM 出力(JSON)を Judgement のリスト(長さ n)へ。欠けた番号は要確認で埋める。"""
    s = raw.strip()
    if s.startswith("```"):
        s = re.sub(r"^```[a-zA-Z0-9_-]*\s*", "", s)
        s = re.sub(r"\s*```$", "", s).strip()
    try:
        data = json.loads(s)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", s, re.DOTALL)
        data = json.loads(m.group(0)) if m else {}

    items = data.get("results") if isinstance(data, dict) else data
    if not isinstance(items, list):
        items = []

    by_no: dict[int, Judgement] = {}
    for it in items:
        if not isinstance(it, dict):
            continue
        try:
            no = int(it.get("no") or it.get("index") or 0)
        except (TypeError, ValueError):
            continue
        if not 1 <= no <= n:
            continue
        mask = _MASK_CANON.get(str(it.get("mask", "")).strip().lower(), "要確認")
        conf = _CONF_CANON.get(str(it.get("confidence", "")).strip().lower(), "低")
        by_no[no] = Judgement(
            mask=mask,
            confidence=conf,
            category=str(it.get("category") or "").strip(),
            reason=str(it.get("reason") or "").strip(),
        )
    return [
        by_no.get(i, _unsure("LLM 応答にこの番号が無かった（要確認へ）"))
        for i in range(1, n + 1)
    ]


# LLM 呼び出しは差し替え可能にする（実機=Azure / テスト=モック）。
# 引数: prompt。戻り: 生の JSON テキスト。
LlmFn = Callable[[str], str]


def _ensure_pii_masker_importable() -> None:
    """pii-masker（pip 非対応の git submodule）を import 可能にする path-injection。

    本体の src/llm/_paths.py と同じ方針で ``external/pii-masker/src`` を sys.path に通す
    （このスクリプトは src に依存しない自己完結にしたいので、その注入だけ写した）。
    submodule 未取得なら分かりやすく落とす。
    """
    src = Path(__file__).resolve().parents[1] / "external" / "pii-masker" / "src"
    if not src.is_dir():
        raise SystemExit(
            f"エラー: pii-masker が見つかりません（{src}）。\n"
            "  git submodule update --init --recursive で取得してください。"
        )
    p = str(src)
    if p not in sys.path:
        sys.path.insert(0, p)


def _azure_llm(model: str) -> LlmFn:
    """pii-masker の Azure クライアントを使う実機用 LLM 呼び出しを返す（遅延 import）。"""
    _ensure_pii_masker_importable()
    from pii_masker.client import get_client  # type: ignore[import-not-found]

    client = get_client(model)

    def call(prompt: str) -> str:
        last: Exception | None = None
        for attempt in range(6):  # 一過性 429 を吸収（指数バックオフ）
            try:
                resp = client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.0,
                    response_format={"type": "json_object"},
                )
                return resp.choices[0].message.content or ""
            except Exception as e:  # noqa: BLE001 - 429 等を粗く再試行
                if "429" not in str(e) and "rate" not in str(e).lower():
                    raise
                last = e
                time.sleep(2.0 * (2**attempt))
        assert last is not None
        raise last

    return call


def _mock_llm(prompt: str) -> str:
    """LLM 無しで配線を検証するための偽物。概要・メモの語で乱暴に振り分けるだけ。

    ⚠ これは判定の**本物ではない**。パース・バッチ・出力の疎通確認専用。
    """
    mask_kw = (
        "自社",
        "独自",
        "専用",
        "社内",
        "コードネーム",
        "プロジェクト",
        "型番",
        "製品名",
    )
    keep_kw = ("一般", "汎用", "標準", "業界", "規格", "共通用語")
    results = []
    # プロンプトから "[i] 用語: ... 概要・メモ: ..." を拾う簡易パース。
    for m in re.finditer(
        r"\[(\d+)\] 用語:.*?(?=\n\n\[\d+\] 用語:|\Z)", prompt, re.DOTALL
    ):
        block = m.group(0)
        no = int(m.group(1))
        memo_m = re.search(r"概要・メモ:\s*(.*)", block, re.DOTALL)
        memo = memo_m.group(1) if memo_m else ""
        if memo.strip() in ("", "(記載なし)"):
            mask, conf, reason = "unsure", "low", "概要・メモが無い(mock)"
        elif any(k in memo for k in mask_kw):
            mask, conf, reason = "yes", "high", "固有性の明確な語を含む(mock)"
        elif any(k in memo for k in keep_kw):
            mask, conf, reason = "no", "mid", "一般語らしい語を含む(mock)"
        else:
            mask, conf, reason = "unsure", "low", "判別語なし(mock)"
        results.append(
            {
                "no": no,
                "mask": mask,
                "confidence": conf,
                "category": "(mock)",
                "reason": reason,
            }
        )
    return json.dumps({"results": results}, ensure_ascii=False)


def classify(
    rows: Sequence[GlossaryRow],
    llm: LlmFn,
    batch_size: int,
    progress: Callable[[int, int], None] | None = None,
    done: dict[int, Judgement] | None = None,
    on_result: Callable[[GlossaryRow, Judgement], None] | None = None,
    throttle: float = 0.0,
) -> list[Judgement]:
    """全行をバッチに分けて LLM 判定し、rows と同順の Judgement 列を返す。

    大量データ対策:
    - ``done``（``{行番号: Judgement}``）にある行は再判定しない（resume 用）。
    - ``on_result`` を各行の判定確定時に呼ぶ（checkpoint への逐次保存用）。
    - ``throttle`` 秒をバッチ間に挟む（429 レート制限の緩和）。
    """
    done = dict(done or {})  # resume 済み（既に checkpoint に入っている）
    results: dict[int, Judgement] = dict(done)  # 返却用（失敗分も含める）
    todo = [r for r in rows if r.row not in done]
    total = (len(todo) + batch_size - 1) // batch_size
    for bi, start in enumerate(range(0, len(todo), batch_size), start=1):
        batch = todo[start : start + batch_size]
        if progress is not None:
            progress(bi, total)
        if throttle > 0:
            time.sleep(throttle)
        try:
            judgements = _parse_results(llm(build_prompt(batch)), len(batch))
            ok = True
        except Exception as e:  # noqa: BLE001 - 1 バッチの失敗で全体を落とさない
            judgements = [_unsure(f"バッチ判定に失敗: {e}") for _ in batch]
            ok = False
        for r, j in zip(batch, judgements):
            results[r.row] = j
            # checkpoint は **成功バッチのみ**。失敗分は残さない＝resume で再試行できる。
            if ok and on_result is not None:
                on_result(r, j)
    return [results.get(r.row, _unsure("未判定")) for r in rows]


# --------------------------------------------------------------------------- #
# チェックポイント（大量データで途中失敗しても再開できるようにする）
# --------------------------------------------------------------------------- #


def _checkpoint_path(out: Path) -> Path:
    """--out に対応する進捗ファイル（JSONL）のパス（例 out.csv → out.progress.jsonl）。"""
    return out.with_name(out.stem + ".progress.jsonl")


def _load_checkpoint(path: Path) -> dict[int, Judgement]:
    """進捗 JSONL を読み、``{行番号: Judgement}`` にする（無ければ空）。壊れた行は飛ばす。"""
    done: dict[int, Judgement] = {}
    if not path.exists():
        return done
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            d = json.loads(line)
        except json.JSONDecodeError:
            continue
        row = d.get("row")
        if isinstance(row, int):
            done[row] = Judgement(
                mask=str(d.get("mask", "要確認")),
                confidence=str(d.get("confidence", "低")),
                category=str(d.get("category", "")),
                reason=str(d.get("reason", "")),
            )
    return done


# --------------------------------------------------------------------------- #
# 出力
# --------------------------------------------------------------------------- #

# 表示・並び順の優先度（要マスク → 要確認 → 不要／確信度 高→低）。
_MASK_ORDER = {"要マスク": 0, "要確認": 1, "不要": 2}
_CONF_ORDER = {"高": 0, "中": 1, "低": 2}


def is_certain(j: Judgement) -> bool:
    """「確実に秘匿すべき」＝要マスク かつ 確信度 高。確定候補の絞り込みに使う。"""
    return j.mask == "要マスク" and j.confidence == "高"


def _sort_key(pair: tuple[GlossaryRow, Judgement]) -> tuple[int, int, str]:
    _, j = pair
    return (_MASK_ORDER.get(j.mask, 9), _CONF_ORDER.get(j.confidence, 9), "")


def write_output(pairs: list[tuple[GlossaryRow, Judgement]], out: Path) -> None:
    rows = [
        {
            "certain": "確実" if is_certain(j) else "",
            "term": r.term,
            "english": r.english or "",
            "japanese": r.japanese or "",
            "kind": r.kind or "",
            "mask": j.mask,
            "confidence": j.confidence,
            "category": j.category,
            "reason": j.reason,
            "memo": r.memo or "",
            "row": r.row,
        }
        for r, j in pairs
    ]
    fields = [
        "certain",
        "term",
        "english",
        "japanese",
        "kind",
        "mask",
        "confidence",
        "category",
        "reason",
        "memo",
        "row",
    ]
    if out.suffix.lower() == ".csv":
        with out.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=fields)
            writer.writeheader()
            writer.writerows(rows)
    else:
        out.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")  # Windows コンソール文字化け対策

    parser = argparse.ArgumentParser(
        description="用語集の各用語を概要・メモから LLM 判定し、秘匿すべき候補を根拠つきで洗い出す。"
    )
    parser.add_argument("glossary", type=Path, help="用語集 Excel(.xlsx/.xlsm)")
    parser.add_argument("--sheet", default=GLOSSARY_SHEET_NAME, help="シート名")
    parser.add_argument(
        "--header-row", type=int, default=GLOSSARY_HEADER_ROW, help="ヘッダ行(1-based)"
    )
    parser.add_argument("--model", default=DEFAULT_MODEL, help="Azure OpenAI モデル")
    parser.add_argument(
        "--batch-size",
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help="1 回に判定する用語数",
    )
    parser.add_argument(
        "--limit", type=int, help="先頭 N 件だけ判定（試用・コスト確認）"
    )
    parser.add_argument(
        "--throttle",
        type=float,
        default=0.5,
        help="バッチ間の待機秒（429 緩和。既定 0.5）",
    )
    parser.add_argument(
        "--no-resume",
        action="store_true",
        help="進捗ファイルがあっても再開せず最初から",
    )
    parser.add_argument(
        "--mock",
        action="store_true",
        help="LLM を呼ばず簡易ヒューリスティックで配線確認",
    )
    parser.add_argument(
        "--out", type=Path, help="結果を書き出す（.csv なら CSV、それ以外は JSON）"
    )
    parser.add_argument(
        "--show-all",
        action="store_true",
        help="画面に全区分を全件表示（既定は★確実のみ＋件数サマリ）",
    )
    parser.add_argument(
        "--print-limit",
        type=int,
        default=50,
        help="★確実の画面表示件数の上限（既定 50。--out には全件出る）",
    )
    args = parser.parse_args(argv)

    if not args.glossary.exists():
        raise SystemExit(f"エラー: 用語集が存在しません: {args.glossary}")

    rows, warnings = load_glossary_rows(args.glossary, args.sheet, args.header_row)
    if args.limit is not None:
        rows = rows[: args.limit]
    print(f"用語集: {args.glossary}（判定対象 {len(rows)} 件）")
    for w in warnings:
        print(f"  ⚠ {w}")

    if args.mock:
        print(
            "  ⚠ --mock: LLM は呼ばず簡易ヒューリスティックで判定（本物ではない・疎通確認用）"
        )
        llm: LlmFn = _mock_llm
    else:
        print(f"  LLM: Azure OpenAI {args.model}（承認済みエンドポイント）")
        llm = _azure_llm(args.model)

    def progress(i: int, n: int) -> None:
        print(f"  判定中… バッチ {i}/{n}", flush=True)

    # 大量データ対策: --out 指定かつ実 LLM のときは進捗を JSONL に逐次保存し、途中失敗を再開できる。
    resume_done: dict[int, Judgement] = {}
    ckpt_fh = None
    ckpt_path: Path | None = None
    if args.out is not None and not args.mock:
        ckpt_path = _checkpoint_path(args.out)
        if not args.no_resume:
            resume_done = _load_checkpoint(ckpt_path)
            if resume_done:
                print(
                    f"  再開: {len(resume_done)} 件は判定済み（{ckpt_path.name}）。残りだけ判定する。"
                )
        ckpt_fh = ckpt_path.open("a", encoding="utf-8")

    def on_result(r: GlossaryRow, j: Judgement) -> None:
        if ckpt_fh is not None:
            ckpt_fh.write(
                json.dumps(
                    {
                        "row": r.row,
                        "mask": j.mask,
                        "confidence": j.confidence,
                        "category": j.category,
                        "reason": j.reason,
                    },
                    ensure_ascii=False,
                )
                + "\n"
            )
            ckpt_fh.flush()  # 落ちても済んだ分が残るよう都度フラッシュ

    throttle = 0.0 if args.mock else args.throttle
    try:
        judgements = classify(
            rows,
            llm,
            args.batch_size,
            progress,
            done=resume_done,
            on_result=on_result,
            throttle=throttle,
        )
    finally:
        if ckpt_fh is not None:
            ckpt_fh.close()
    pairs = sorted(zip(rows, judgements), key=_sort_key)

    def show(pair: tuple[GlossaryRow, Judgement]) -> None:
        r, j = pair
        extra = " / ".join(x for x in (r.english, r.japanese, r.kind) if x)
        suffix = f"  [{extra}]" if extra else ""
        print(f"・[{j.mask}/確信度{j.confidence}] {r.term}{suffix}")
        if j.category:
            print(f"    分類: {j.category}")
        print(f"    根拠: {j.reason or '(なし)'}")

    certain = [p for p in pairs if is_certain(p[1])]
    maybe = [p for p in pairs if p[1].mask == "要マスク" and not is_certain(p[1])]
    review = [p for p in pairs if p[1].mask == "要確認"]
    keep = [p for p in pairs if p[1].mask == "不要"]

    print(
        f"\n判定内訳: 確実に秘匿 {len(certain)} / 秘匿の可能性 {len(maybe)}"
        f" / 要確認 {len(review)} / 不要 {len(keep)}"
    )

    if args.show_all:
        # 全区分を全件表示（従来挙動）。件数が少ないとき・画面で全部見たいとき用。
        sections = [
            (f"★確実に秘匿すべき（要マスク×確信度高・{len(certain)} 件）", certain),
            (f"秘匿の可能性（要マスク×中/低・{len(maybe)} 件・要レビュー）", maybe),
            (f"要確認（判断材料不足・{len(review)} 件）", review),
            (f"不要（{len(keep)} 件）", keep),
        ]
        for title, group in sections:
            print(f"\n===== {title} =====")
            for p in group:
                show(p)
            if not group:
                print("  （該当なし）")
    else:
        # 既定は画面を汚さない: 成果物の核＝★確実だけ出す。残りは件数のみ（詳細は --out へ）。
        cap = args.print_limit
        print(
            f"\n===== ★確実に秘匿すべき（要マスク×確信度高・{len(certain)} 件） ====="
        )
        if certain:
            for p in certain[:cap]:
                show(p)
            if len(certain) > cap:
                print(f"  … 他 {len(certain) - cap} 件（全件は --out / --show-all で）")
        else:
            print("  （該当なし）")
        print(
            f"\n秘匿の可能性 {len(maybe)} 件 / 要確認 {len(review)} 件 / 不要 {len(keep)} 件"
            " は画面では省略。"
        )
        if args.out is not None:
            print("  → 全件の詳細（根拠つき）は書き出したファイルを参照。")
        else:
            print(
                "  → 全件を残すには --out out.csv を付ける（--show-all で画面に全表示）。"
            )

    if args.out is not None:
        write_output(pairs, args.out)
        print(f"\n結果を書き出しました: {args.out.resolve()}")
        # 失敗バッチが無ければ進捗ファイルは不要。残っていれば再実行で残りを再開できる。
        if ckpt_path is not None and ckpt_path.exists():
            had_failure = any("バッチ判定に失敗" in j.reason for _, j in pairs)
            if had_failure:
                print(
                    f"  ⚠ 失敗バッチあり。進捗 {ckpt_path.name} を残す（同じ --out で再実行すれば残りだけ判定）。"
                )
            else:
                ckpt_path.unlink()
    return 0


if __name__ == "__main__":
    sys.exit(main())
