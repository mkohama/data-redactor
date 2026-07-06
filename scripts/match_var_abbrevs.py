"""変数名から拾った略語候補（Excel の B 列）を用語集と突き合わせ、**大小無視で一致**する語を抽出する
使い捨てスクリプト。

用途:
  ソースの変数名から機械的に拾った略語らしき語（例 ``Nsr`` ``Cmp``）が並ぶ Excel と、用語集
  （``用語`` 列を持つ Excel）を突き合わせ、**大文字・小文字を無視して一致**する候補だけを残す。
  ＝「変数に出てくる略語のうち、用語集に載っている（＝社内で定義された）もの」を洗い出す。
  マスク辞書の候補づくりの前段（人手レビュー前のふるい）。

前提フォーマット:
  - 候補 Excel: **B 列**に略語候補（1 行 1 語）。既定でシート先頭・1 行目はヘッダとして読み飛ばす
    （``--cand-col`` / ``--cand-sheet`` / ``--cand-header-row`` で変更。ヘッダ無しなら ``--cand-header-row 0``）。
  - 用語集 Excel: シート「用語集」・6 行目ヘッダ・列「用語 / 英語名称 / 日本語名称」（列は**ヘッダ名で探す**
    ので位置がずれても可。``--glossary-sheet`` / ``--glossary-header-row`` で変更）。find_feature_names /
    classify_glossary_terms と同じ形式。

一致の定義:
  NFKC 正規化＋casefold（大小・全角半角を無視）した完全一致。``Nsr`` は用語集 ``NSR`` に一致。
  用語集側はカンマ区切り（``BUF1,BUF2``）を分割し、括弧内の別表記（``ALGO-M (AM)`` の ``AM``）も
  一致対象に含める（find_feature_names と同じ前処理）。部分一致はしない（``Nsr`` は ``NsrCtrl`` に当てない）。

  **既定では「表記が異なるもの」だけを出す**：候補の綴りが用語集にそのまま在る＝完全一致
  （``NSR`` と ``NSR``）は除外し、大小・全半角などの**表記差**だけを持つもの（``Nsr`` ↔ ``NSR``）を残す。
  ＝「用語集と綴りが揺れている変数略語」を洗い出す用途。全件見たいときは ``--include-exact``。

使い方:
    uv run python scripts/match_var_abbrevs.py <候補.xlsx> <用語集.xlsx>
    uv run python scripts/match_var_abbrevs.py <候補.xlsx> <用語集.xlsx> --out hits.csv
    # 候補列やヘッダ位置を変える
    uv run python scripts/match_var_abbrevs.py <候補.xlsx> <用語集.xlsx> \
        --cand-col B --cand-sheet Sheet1 --cand-header-row 1 \
        --glossary-sheet 用語集 --glossary-header-row 6
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path

# --------------------------------------------------------------------------- #
# 正規化（大小・全角半角・空白を無視。src/masking の normalize と同方針）
# --------------------------------------------------------------------------- #

_WHITESPACE = re.compile(r"\s+")


def normalize(text: str) -> str:
    """照合キー：NFKC で全角半角統一・casefold で大小無視・空白除去。"""
    return _WHITESPACE.sub("", unicodedata.normalize("NFKC", text).casefold())


# --------------------------------------------------------------------------- #
# 用語集パース（find_feature_names / classify_glossary_terms と同形式）
# --------------------------------------------------------------------------- #

GLOSSARY_SHEET_NAME = "用語集"
GLOSSARY_HEADER_ROW = 6
COL_TERM = "用語"
COL_EN = "英語名称"
COL_JA = "日本語名称"
_GLOSSARY_COLS = (COL_TERM, COL_EN, COL_JA)
_HEADER_SEARCH_LIMIT = 30  # 指定行にヘッダが無いときに走査する行数

# 用語の正規化補助：括弧内の別表記を取り出す（ALGO-M (AM) → AM も一致対象に）。
_PAREN_RE = re.compile(r"[（(]([^）)]*)[）)]")


def _cell(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


@dataclass
class GlossaryTerm:
    """用語集 1 語（照合キーは正規化して別途 index する）。"""

    raw: str  # 用語列の原文（表示用）
    english: str | None
    japanese: str | None
    row: int


def _find_header(
    sheet, header_row: int, cols: tuple[str, ...]
) -> tuple[int, dict] | None:
    """ヘッダ行番号と列名→インデックスの対応を返す。指定行→先頭数十行の順に探索。"""
    candidates = [header_row] + [
        r for r in range(1, _HEADER_SEARCH_LIMIT + 1) if r != header_row
    ]
    for r in candidates:
        row = next(sheet.iter_rows(min_row=r, max_row=r, values_only=True), None)
        if row is None:
            continue
        header_map: dict[str, int] = {}
        for idx, cell in enumerate(row):
            name = _cell(cell)
            if name in cols and name not in header_map:
                header_map[name] = idx
        if COL_TERM in header_map:
            return r, header_map
    return None


def _term_pieces(raw: str) -> list[tuple[str, str]]:
    """用語 1 セルから (正規化キー, 原文ピース) の列を作る。カンマ分割＋括弧内別表記も対象に。

    原文ピース（原表記のまま）を持つのは、候補と**同じ綴りか否か**（完全一致か表記差か）を
    判定するため。例: 用語 ``ALGO-M (AM)`` → (``algom``,``ALGO-M``) と (``am``,``AM``)。
    """
    out: list[tuple[str, str]] = []
    # 括弧内（別表記）を取り出しつつ、括弧を除いた本体も対象にする。
    inners = _PAREN_RE.findall(raw)
    body = _PAREN_RE.sub(" ", raw)
    for chunk in [body, *inners]:
        for piece in re.split(r"[,、，]", chunk):  # カンマ区切りを分割
            p = piece.strip()
            key = normalize(p)
            if key:
                out.append((key, p))
    return out


def load_glossary(
    path: Path, sheet_name: str, header_row: int
) -> tuple[dict[str, list[GlossaryTerm]], dict[str, set[str]], list[str]]:
    """用語集を読み、(正規化キー→語 index, 正規化キー→原文ピース集合, 警告) を返す。

    原文ピース集合は「候補と同じ綴りか」を見るため（完全一致の除外に使う）。
    """
    from openpyxl import load_workbook  # type: ignore[import-untyped]  # スタブ無し

    warnings: list[str] = []
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"エラー: シート {sheet_name!r} が見つかりません（存在: {wb.sheetnames}）"
        )
    sheet = wb[sheet_name]

    found = _find_header(sheet, header_row, _GLOSSARY_COLS)
    if found is None:
        raise SystemExit(
            f"エラー: ヘッダ行に {COL_TERM!r} 列が見つかりません"
            f"（{header_row} 行目と先頭 {_HEADER_SEARCH_LIMIT} 行を確認）"
        )
    actual_row, header_map = found
    if actual_row != header_row:
        warnings.append(f"ヘッダを {header_row} 行目でなく {actual_row} 行目で検出。")
    for col in (COL_EN, COL_JA):
        if col not in header_map:
            warnings.append(f"{col!r} 列なし（その情報なしで続行）。")

    ti = header_map[COL_TERM]
    ei = header_map.get(COL_EN)
    ji = header_map.get(COL_JA)

    def pick(row: tuple, idx: int | None) -> str | None:
        return _cell(row[idx]) if idx is not None and len(row) > idx else None

    index: dict[str, list[GlossaryTerm]] = {}
    forms: dict[str, set[str]] = {}  # 正規化キー → 原文ピース集合（表記差判定用）
    for n, row in enumerate(
        sheet.iter_rows(min_row=actual_row + 1, values_only=True), start=actual_row + 1
    ):
        raw = _cell(row[ti]) if len(row) > ti else None
        if raw is None:
            continue
        term = GlossaryTerm(
            raw=raw, english=pick(row, ei), japanese=pick(row, ji), row=n
        )
        for key, piece in _term_pieces(raw):
            index.setdefault(key, []).append(term)
            forms.setdefault(key, set()).add(piece)
    return index, forms, warnings


# --------------------------------------------------------------------------- #
# 候補（B 列）読み込み
# --------------------------------------------------------------------------- #


def _col_to_index(letter: str) -> int:
    from openpyxl.utils import column_index_from_string  # type: ignore[import-untyped]

    return column_index_from_string(letter.strip().upper()) - 1


@dataclass
class Candidate:
    """候補 1 語（原文と行番号。照合は正規化して行う）。"""

    raw: str
    row: int


def load_candidates(
    path: Path, sheet: str | None, header_row: int, col: str
) -> list[Candidate]:
    """候補 Excel の指定列（既定 B）から語を読む（空白・重複は除く。原文の初出を残す）。"""
    from openpyxl import load_workbook  # type: ignore[import-untyped]  # スタブ無し

    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet is not None:
        if sheet not in wb.sheetnames:
            raise SystemExit(
                f"エラー: シート {sheet!r} が見つかりません（存在: {wb.sheetnames}）"
            )
        ws = wb[sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    ci = _col_to_index(col)
    out: list[Candidate] = []
    seen: set[str] = set()
    for n, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        raw = _cell(row[ci]) if len(row) > ci else None
        if raw is None or not normalize(raw):
            continue
        # 重複除去は**原表記が同一**のときだけ（Nsr と NSR は別綴り＝別候補として残す。
        # 大小で畳むと、綴り違いの片方を取りこぼす）。
        if raw in seen:
            continue
        seen.add(raw)
        out.append(Candidate(raw=raw, row=n))
    return out


# --------------------------------------------------------------------------- #
# 突き合わせ・出力
# --------------------------------------------------------------------------- #


@dataclass
class Hit:
    """一致 1 件（候補 → 用語集の語）。"""

    candidate: str
    glossary_terms: list[str]
    english: list[str]
    japanese: list[str]


def match(
    candidates: list[Candidate],
    glossary: dict[str, list[GlossaryTerm]],
    forms: dict[str, set[str]],
    include_exact: bool = False,
) -> tuple[list[Hit], list[Candidate], int]:
    """候補を用語集キーに突き合わせ、(一致, 非一致, 完全一致で除外した数) を返す。

    既定では**表記が異なるもの（大小・全半角等の差）だけ**を一致として返す。候補の原文が
    用語集の原文ピースにそのまま在る＝**完全一致**は除外する（``include_exact=True`` で含める）。
    """
    hits: list[Hit] = []
    misses: list[Candidate] = []
    exact_excluded = 0
    for c in candidates:
        key = normalize(c.raw)
        terms = glossary.get(key)
        if not terms:
            misses.append(c)
            continue
        # 候補の原文がそのまま用語集にある＝綴りが同じ＝表記差なし → 既定では除外。
        if not include_exact and c.raw in forms.get(key, set()):
            exact_excluded += 1
            continue
        # 同じキーに複数の用語行が当たり得る（別表記の重複等）。重複を畳んで併記。
        raws = list(dict.fromkeys(t.raw for t in terms))
        ens = list(dict.fromkeys(t.english for t in terms if t.english))
        jas = list(dict.fromkeys(t.japanese for t in terms if t.japanese))
        hits.append(
            Hit(candidate=c.raw, glossary_terms=raws, english=ens, japanese=jas)
        )
    return hits, misses, exact_excluded


_CSV_FIELDS = ["candidate", "glossary_terms", "english", "japanese"]


def write_output(hits: list[Hit], out: Path) -> None:
    rows = [
        {
            "candidate": h.candidate,
            "glossary_terms": " / ".join(h.glossary_terms),
            "english": " / ".join(h.english),
            "japanese": " / ".join(h.japanese),
        }
        for h in hits
    ]
    if out.suffix.lower() == ".csv":
        with out.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=_CSV_FIELDS)
            writer.writeheader()
            writer.writerows(rows)
    else:
        out.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")  # Windows コンソール文字化け対策

    parser = argparse.ArgumentParser(
        description="変数名由来の略語候補（B列）を用語集と大小無視で突き合わせ、一致を抽出する。"
    )
    parser.add_argument("candidates", type=Path, help="候補 Excel（B列に略語候補）")
    parser.add_argument("glossary", type=Path, help="用語集 Excel（用語列を持つ）")
    parser.add_argument("--cand-col", default="B", help="候補の列（既定 B）")
    parser.add_argument("--cand-sheet", default=None, help="候補シート名（既定: 先頭）")
    parser.add_argument(
        "--cand-header-row",
        type=int,
        default=1,
        help="候補のヘッダ行(1-based。既定 1。ヘッダ無しなら 0)",
    )
    parser.add_argument(
        "--glossary-sheet", default=GLOSSARY_SHEET_NAME, help="用語集シート名"
    )
    parser.add_argument(
        "--glossary-header-row",
        type=int,
        default=GLOSSARY_HEADER_ROW,
        help="用語集ヘッダ行(1-based。既定 6)",
    )
    parser.add_argument(
        "--include-exact",
        action="store_true",
        help="完全一致（綴りが用語集と同一）も含める（既定は表記が異なるものだけ）",
    )
    parser.add_argument(
        "--out", type=Path, help="結果を書き出す（.csv なら CSV、それ以外は JSON）"
    )
    args = parser.parse_args(argv)

    for p in (args.candidates, args.glossary):
        if not p.exists():
            raise SystemExit(f"エラー: ファイルが存在しません: {p}")

    glossary, forms, g_warn = load_glossary(
        args.glossary, args.glossary_sheet, args.glossary_header_row
    )
    for w in g_warn:
        print(f"  ⚠ 用語集: {w}")
    candidates = load_candidates(
        args.candidates, args.cand_sheet, args.cand_header_row, args.cand_col
    )
    print(
        f"候補: {len(candidates)} 語（重複除去後） / 用語集: {len(glossary)} キー"
        f"（カンマ・括弧展開後）"
    )

    hits, misses, exact_excluded = match(
        candidates, glossary, forms, include_exact=args.include_exact
    )
    tail = "" if args.include_exact else f" / 完全一致で除外 {exact_excluded} 語"
    print(f"\n一致: {len(hits)} 語 / 非一致: {len(misses)} 語{tail}\n")

    label = (
        "用語集と一致（すべて）"
        if args.include_exact
        else "用語集と表記が異なる略語候補"
    )
    print(f"===== {label} =====")
    if hits:
        for h in hits:
            extra = " / ".join(
                x for x in (" ".join(h.english), " ".join(h.japanese)) if x
            )
            terms = " / ".join(h.glossary_terms)
            suffix = f"  [{extra}]" if extra else ""
            print(f"・{h.candidate}  → 用語集: {terms}{suffix}")
    else:
        print("  （該当なし）")

    if args.out is not None:
        write_output(hits, args.out)
        print(f"\n一致を書き出しました: {args.out.resolve()}")
    else:
        print("\n（--out out.csv で一致一覧を書き出せる）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
