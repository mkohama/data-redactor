"""用語集(Excel)の用語のうち、指定フォルダ配下のファイル名に現れるものを
「マスク対象候補の機能名」として洗い出す使い捨てスクリプト。

やること:
  (1) 用語集 Excel(.xlsx/.xlsm)を解析し「用語」列を集める
      （term-variants の import_glossary.py / glossary.py と同じフォーマット前提:
       シート「用語集」・6 行目ヘッダ・「用語 / 英語名称 / 日本語名称」列）。
      照合キーは「用語」列のみ。英語名称・日本語名称はヒット時の参考情報。
  (2) 指定フォルダ配下のファイル名を走査し、各用語がファイル名に含まれるか単語境界一致で判定。
      ヒットした用語＝とりあえずのマスク候補となる機能名。

マッチ方式（単語境界一致）:
  - ラテン文字主体の用語: ファイル名を識別子分割（区切り文字・camelCase・大文字連続・
    英数境界）したトークン列に、用語のトークン列が「連続部分列」として現れるときだけヒット。
    → AM が PARAM / CAMERA に部分一致する誤検出を防ぐ。
  - CJK(かな/漢字/カナ)・記号を含む用語: NFKC 正規化したファイル名への部分文字列一致。

使い方:
  uv run python scripts/find_feature_names.py <用語集.xlsm> <対象フォルダ> [--out out.csv]
  # シート名/ヘッダ行が違う場合:  --sheet 用語集 --header-row 6
  # ディレクトリ名も見る:         --include-dirs
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

# --------------------------------------------------------------------------- #
# 用語集(Excel)パース
# --------------------------------------------------------------------------- #

GLOSSARY_SHEET_NAME = "用語集"
GLOSSARY_HEADER_ROW = 6
COL_TERM = "用語"
COL_EN = "英語名称"
COL_JA = "日本語名称"
_HEADER_SEARCH_LIMIT = 30  # 指定行にヘッダが無いときに走査する行数

_WILDCARD_RE = re.compile(r"\*\*")
_BRACKET_RE = re.compile(r"\[[^\]]*\]")
_IDENT_PART_RE = re.compile(r"^[A-Za-z0-9_\-]+$")
_PAREN_CLOSED_RE = re.compile(r"[(（][^)）]*[)）]")
_PAREN_OPEN_RE = re.compile(r"[(（]")


@dataclass(frozen=True)
class TermEntry:
    term: str  # 照合キー（NFKC 正規化・括弧除去後の「用語」）
    raw_term: str  # 「用語」列の原文
    english: str | None
    japanese: str | None
    row: int


def _cell(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def normalize_term(surface: str) -> str:
    """NFKC 正規化 + 括弧除去 + 空白圧縮（例: 'ALGO-M (AM)' -> 'ALGO-M'）。"""
    cleaned = unicodedata.normalize("NFKC", surface)
    cleaned = _PAREN_CLOSED_RE.sub("", cleaned)
    m = _PAREN_OPEN_RE.search(cleaned)
    if m:
        cleaned = cleaned[: m.start()]
    return re.sub(r"\s+", " ", cleaned).strip()


def _expand_terms(raw_term: str, warnings: list[str], row: int) -> list[str]:
    """1 つの「用語」セルを照合キー（複数可）に展開する。"""
    if _WILDCARD_RE.search(raw_term) and _BRACKET_RE.search(raw_term):
        warnings.append(
            f"行 {row}: ワイルドカード/ブラケットを含む用語は自動展開不可"
            f"（原文のまま照合）: {raw_term!r}"
        )
        return [raw_term.strip()]
    if "," in raw_term:
        parts = [p.strip() for p in raw_term.split(",")]
        if len(parts) >= 2 and all(p and _IDENT_PART_RE.match(p) for p in parts):
            return parts
    cleaned = normalize_term(raw_term)
    return [cleaned] if cleaned else []


def _find_header(sheet, header_row: int) -> tuple[int, dict[str, int]] | None:
    """ヘッダ行番号と列インデックスの対応を返す。まず指定行、無ければ先頭数十行を探索。"""
    candidates = [header_row] + [
        r for r in range(1, _HEADER_SEARCH_LIMIT + 1) if r != header_row
    ]
    for r in candidates:
        row = next(sheet.iter_rows(min_row=r, max_row=r, values_only=True), None)
        if row is None:
            continue
        header_map: dict[str, int] = {}
        for idx, cell in enumerate(row):
            name = _cell(cell)
            if name in (COL_TERM, COL_EN, COL_JA) and name not in header_map:
                header_map[name] = idx
        if COL_TERM in header_map:
            return r, header_map
    return None


def load_glossary_terms(
    path: Path, sheet_name: str, header_row: int
) -> tuple[list[TermEntry], list[str]]:
    """用語集 Excel から TermEntry 群を読む。「用語」列さえあれば読める。"""
    from openpyxl import load_workbook  # type: ignore[import-untyped]  # スタブ無し

    warnings: list[str] = []
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"エラー: シート {sheet_name!r} が見つかりません（存在: {wb.sheetnames}）"
        )
    sheet = wb[sheet_name]

    found = _find_header(sheet, header_row)
    if found is None:
        raise SystemExit(
            f"エラー: ヘッダ行に {COL_TERM!r} 列が見つかりません"
            f"（{header_row} 行目と先頭 {_HEADER_SEARCH_LIMIT} 行を確認）"
        )
    actual_row, header_map = found
    if actual_row != header_row:
        warnings.append(f"ヘッダを {header_row} 行目でなく {actual_row} 行目で検出。")
    for col in (COL_EN, COL_JA):
        if col not in header_map:
            warnings.append(f"{col!r} 列なし（参考情報なしで続行）。")

    ti = header_map[COL_TERM]
    ei = header_map.get(COL_EN)
    ji = header_map.get(COL_JA)

    entries: list[TermEntry] = []
    for n, row in enumerate(
        sheet.iter_rows(min_row=actual_row + 1, values_only=True), start=actual_row + 1
    ):
        raw_term = _cell(row[ti]) if len(row) > ti else None
        if raw_term is None:
            continue
        english = _cell(row[ei]) if ei is not None and len(row) > ei else None
        japanese = _cell(row[ji]) if ji is not None and len(row) > ji else None
        for key in _expand_terms(raw_term, warnings, n):
            entries.append(TermEntry(key, raw_term, english, japanese, n))
    return entries, warnings


# --------------------------------------------------------------------------- #
# ファイル名との照合（単語境界一致）
# --------------------------------------------------------------------------- #

# 識別子分割: 大文字連続(AM/HTTP) / Titlecase 語(Value) / 小文字語 / 数字。区切り文字は非マッチ。
_TOKEN_RE = re.compile(r"[A-Z]+(?![a-z])|[A-Z][a-z]+|[a-z]+|[0-9]+")
_CJK_RE = re.compile(r"[぀-ヿ㐀-䶿一-鿿ｦ-ﾟ]")


def tokenize_ascii(s: str) -> list[str]:
    return [m.group(0).lower() for m in _TOKEN_RE.finditer(s)]


def _subseq(haystack: list[str], needle: list[str]) -> bool:
    n, m = len(haystack), len(needle)
    if m == 0 or m > n:
        return False
    first = needle[0]
    for i in range(n - m + 1):
        if haystack[i] == first and haystack[i : i + m] == needle:
            return True
    return False


@dataclass(frozen=True)
class _Matcher:
    is_token: bool
    tokens: tuple[str, ...]
    substr: str


def build_matcher(term: str) -> _Matcher | None:
    if not term:
        return None
    tokens = tokenize_ascii(term)
    if tokens and not _CJK_RE.search(term):
        return _Matcher(True, tuple(tokens), "")
    substr = unicodedata.normalize("NFKC", term).casefold()
    return _Matcher(False, (), substr) if substr else None


@dataclass
class _FileKey:
    tokens: list[str]
    token_set: frozenset[str]
    substr: str


def build_file_key(stem: str) -> _FileKey:
    tokens = tokenize_ascii(stem)
    return _FileKey(
        tokens, frozenset(tokens), unicodedata.normalize("NFKC", stem).casefold()
    )


def matches(matcher: _Matcher, key: _FileKey) -> bool:
    if matcher.is_token:
        if len(matcher.tokens) == 1:
            return matcher.tokens[0] in key.token_set
        return _subseq(key.tokens, list(matcher.tokens))
    return matcher.substr in key.substr


# --------------------------------------------------------------------------- #
# フォルダ走査
# --------------------------------------------------------------------------- #


@dataclass
class TermHit:
    term: str
    raw_term: str
    english: str | None
    japanese: str | None
    files: list[Path] = field(default_factory=list)


def scan_folder(
    entries: list[TermEntry],
    folder: Path,
    include_dirs: bool,
) -> tuple[list[TermHit], int, int]:
    """フォルダ配下の名前を走査し (hits, 走査数, 照合不能な用語数) を返す。"""
    matchers: list[tuple[TermEntry, _Matcher]] = []
    skipped = 0
    for e in entries:
        m = build_matcher(e.term)
        if m is None:
            skipped += 1
        else:
            matchers.append((e, m))

    seen: dict[str, TermHit] = {}
    order: list[str] = []
    scanned = 0
    for p in folder.rglob("*"):
        is_file = p.is_file()
        if p.is_dir():
            if not include_dirs:
                continue
        elif not is_file:
            continue
        scanned += 1
        name = p.stem if is_file else p.name  # 拡張子は落とす
        key = build_file_key(name)
        try:
            rel = p.relative_to(folder)
        except ValueError:
            rel = p
        for e, m in matchers:
            if matches(m, key):
                hit = seen.get(e.term)
                if hit is None:
                    hit = TermHit(e.term, e.raw_term, e.english, e.japanese)
                    seen[e.term] = hit
                    order.append(e.term)
                hit.files.append(rel)

    hits = [seen[t] for t in order]
    hits.sort(key=lambda h: (-len(h.files), h.term))
    return hits, scanned, skipped


# --------------------------------------------------------------------------- #
# 出力
# --------------------------------------------------------------------------- #


def write_output(hits: list[TermHit], out: Path) -> None:
    rows = [
        {
            "term": h.term,
            "raw_term": h.raw_term,
            "english": h.english,
            "japanese": h.japanese,
            "file_count": len(h.files),
            "files": [str(f) for f in h.files],
        }
        for h in hits
    ]
    if out.suffix.lower() == ".csv":
        with out.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(
                ["term", "raw_term", "english", "japanese", "file_count", "files"]
            )
            for h in hits:  # CSV は hits から直接（rows の混在型 union を通さない）
                writer.writerow(
                    [
                        h.term,
                        h.raw_term,
                        h.english or "",
                        h.japanese or "",
                        len(h.files),
                        " | ".join(str(f) for f in h.files),
                    ]
                )
    else:
        out.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")  # Windows コンソール文字化け対策

    parser = argparse.ArgumentParser(
        description="用語集の用語がファイル名に現れるものを機能名候補として洗い出す。"
    )
    parser.add_argument("glossary", type=Path, help="用語集 Excel(.xlsx/.xlsm)")
    parser.add_argument("folder", type=Path, help="走査対象フォルダ（再帰）")
    parser.add_argument("--sheet", default=GLOSSARY_SHEET_NAME, help="シート名")
    parser.add_argument(
        "--header-row", type=int, default=GLOSSARY_HEADER_ROW, help="ヘッダ行(1-based)"
    )
    parser.add_argument(
        "--include-dirs", action="store_true", help="ディレクトリ名も照合対象にする"
    )
    parser.add_argument(
        "--out", type=Path, help="結果を書き出す（.csv なら CSV、それ以外は JSON）"
    )
    args = parser.parse_args(argv)

    if not args.glossary.exists():
        raise SystemExit(f"エラー: 用語集が存在しません: {args.glossary}")
    if not args.folder.is_dir():
        raise SystemExit(f"エラー: フォルダが存在しません: {args.folder}")

    entries, warnings = load_glossary_terms(args.glossary, args.sheet, args.header_row)
    print(f"用語集: {args.glossary}（照合キー {len(entries)} 件）")
    for w in warnings:
        print(f"  ⚠ {w}")

    print(f"走査: {args.folder}（配下のファイル名を単語境界一致）")
    hits, scanned, skipped = scan_folder(entries, args.folder, args.include_dirs)
    print(f"走査対象 {scanned} 件 / 照合できなかった用語 {skipped} 件")

    print(f"\n===== 機能名候補（{len(hits)} 件・ファイル数の多い順） =====")
    for h in hits:
        extra = " / ".join(x for x in (h.english, h.japanese) if x)
        suffix = f"  [{extra}]" if extra else ""
        print(f"・{h.term}\tファイル{len(h.files)}件{suffix}")
        for f in h.files:
            print(f"    {f}")

    if args.out is not None:
        write_output(hits, args.out)
        print(f"\n候補を書き出しました: {args.out.resolve()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
