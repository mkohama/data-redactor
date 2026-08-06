"""用語集と現行のマスク辞書を突き合わせ、各エントリが用語集由来かどうかを一覧にする。

なぜ必要か:
  用語集から一括登録した語にノイズが多く、それだけを取り消したいという要望がある。
  ところが辞書 YAML には出所を示す項目が無く、data/*.yaml は git 管理外で履歴も残らない。
  そこで「用語集に載っている語かどうか」を後から突き合わせて判定する。

判定のしかた:
  照合キーは src/masking/dictionary.py の normalize（NFKC＋casefold＋空白除去）。辞書の照合と
  同じ正規化なので、大小・全角半角・空白の違いは吸収される。辞書エントリは代表表記だけでなく
  別名でも突き合わせる（別名が用語集にあれば、その語は用語集由来とみなす）。

  用語集側の入力は 2 通りを受ける:
    - Excel（.xlsx）… scripts/build_mask_dict.py と同じ読み取り（列位置・ヘッダ行）を使う。
      同じ関数を呼ぶので「登録が必要な用語」の取り出し方が build_mask_dict と必ず一致する。
    - YAML（.yaml/.yml）… build_mask_dict.py --out で書き出した辞書 YAML。

このスクリプトは**読むだけ**で、辞書は書き換えない。削除するかどうかは一覧を見て人が決める。

使い方:
    # まず用語集の中身を見て、用語がどの列にあるかを確かめる（0 語になったときはこれ）
    uv run python scripts/compare_glossary.py 用語集.xlsx --inspect

    # Excel の用語集と現行辞書を突き合わせる（列位置の既定は build_mask_dict と同じ）
    uv run python scripts/compare_glossary.py 用語集.xlsx

    # build_mask_dict.py が書き出した YAML と突き合わせる
    uv run python scripts/compare_glossary.py glossary.yaml --dict data/mask_dict.yaml

    # 用語集由来だけを CSV で（Excel で開いてレビューする用）
    uv run python scripts/compare_glossary.py 用語集.xlsx --only glossary --csv \
        --out glossary_origin.csv

    # 列位置・ヘッダ行を指定する（用語の列は複数可）。
    # 例: C列=用語・D列=英語名称・E列=日本語名称、6行目がヘッダの用語集
    uv run python scripts/compare_glossary.py 用語集.xlsx --term-col C,D,E --header-row 6

前提となる Excel の形は 2 種類あるので注意する:
  - **元の用語集**… 用語が並んだ一覧。列位置は用語集ごとに違うので --term-col / --header-row /
    --sheet で指定する（--inspect で下見できる）。
  - **キュレーション済み Excel**… build_mask_dict.py に渡して辞書 YAML を作るための表
    （A列=用語 / D列=種類 / E列=部分一致 / F列=大小区別・1行目ヘッダ）。既定値はこちらに合わせてある。
"""

from __future__ import annotations

import argparse
import csv
import io
import sys
from pathlib import Path

# scripts/ はパッケージではないので、同じディレクトリを import パスに足して build_mask_dict を使う
# （用語の取り出し方を 1 か所に保つ＝取りこぼしの仕様差を作らない）。
sys.path.insert(0, str(Path(__file__).resolve().parent))
# リポジトリルート（src を import するため）。
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.masking.dictionary import load_entries, normalize  # noqa: E402


def glossary_terms(args: argparse.Namespace) -> dict[str, tuple[str, str]]:
    """用語集の語を ``{正規化キー: (元の表記, 出どころの説明)}`` で返す。

    出どころの説明は Excel なら ``行12``、YAML なら ``Trademark`` のようにその語を見つけた
    位置。一覧に添えて、辞書のどのエントリがどの行に対応するかを追えるようにする。
    """
    path = Path(args.glossary)
    if not path.exists():
        raise SystemExit(f"エラー: 用語集が見つかりません: {path}")

    out: dict[str, tuple[str, str]] = {}
    if path.suffix.lower() in (".yaml", ".yml"):
        for e in load_entries(path):
            for surface in [e["canonical"], *e["aliases"]]:
                if surface:
                    out.setdefault(normalize(surface), (surface, e["category"]))
        return out

    # Excel は build_mask_dict と同じ読み取りを使う（列位置・ヘッダ行の解釈まで共通）。
    import build_mask_dict as bmd  # type: ignore[import-not-found]

    # 用語の列は複数指定できる（例 --term-col C,D,E＝用語・英語名称・日本語名称）。
    #   1 つの概念が複数の表記で載っている用語集があり、辞書にどの表記で登録されているかは
    #   分からないため、指定された全列を突き合わせ対象の表記として集める。
    #   列ごとに load_rows を呼ぶ＝読み取りの解釈は build_mask_dict と 1 本のまま。
    for column in [c.strip() for c in args.term_col.split(",") if c.strip()]:
        rows = bmd.load_rows(
            path,
            args.sheet,
            args.header_row,
            column,
            args.kind_col,
            args.partial_col,
            args.case_col,
        )
        for raw in rows:
            term = (raw.term or "").strip()
            if term:
                out.setdefault(normalize(term), (term, f"{column}{raw.row}"))
    return out


def inspect_glossary(
    path: Path, sheet: str | None, header_row: int, rows: int, cols: int
) -> str:
    """用語集 Excel の中身を列記号つきで表示する（列位置を特定するための下見）。

    「用語集 0 語」になったときに、どの列に用語が入っているか・ヘッダが何行目か・シートが
    どれかを目で確かめるために使う。--term-col などに何を渡せばよいかはこれを見て決める。
    セルは data_only=True で読む（build_mask_dict と同じ）。数式しか無く計算結果が保存されて
    いないファイルは、ここでも空に見える＝それが 0 語の原因と分かる。
    """
    from openpyxl import load_workbook  # type: ignore[import-untyped]  # スタブ無し
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

    wb = load_workbook(path, data_only=True, read_only=True)
    out = [f"ファイル: {path}", "シート一覧（--sheet で選べる）:"]
    for name in wb.sheetnames:
        s = wb[name]
        size = f"{s.max_row or '?'} 行 x {s.max_column or '?'} 列"
        out.append(f"  {name}  ({size})")
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    out.append("")
    out.append(
        f"読むシート: {ws.title}" + ("" if sheet else "（--sheet 未指定＝先頭）")
    )
    out.append(f"ヘッダ行の指定: {header_row}（{header_row + 1} 行目から読む）")
    out.append("")

    for n, row in enumerate(ws.iter_rows(max_row=rows, values_only=True), start=1):
        marker = "ヘッダ" if n <= header_row else "データ"
        out.append(f"[{n:>3} 行目 / {marker}]")
        for i, value in enumerate(row[:cols]):
            if value is None or str(value).strip() == "":
                continue
            text = " ".join(str(value).split())
            out.append(f"    {get_column_letter(i + 1)} : {text[:60]}")
    out.append("")
    out.append(
        "用語が入っている列の記号を --term-col に、ヘッダの行数を --header-row に指定する。"
    )
    out.append("例: 用語が B 列・ヘッダが 3 行目なら  --term-col B --header-row 3")
    return "\n".join(out) + "\n"


def classify(dict_path: Path, terms: dict[str, tuple[str, str]]) -> list[dict]:
    """辞書の各エントリに「用語集由来か」を付けて返す。

    代表表記が用語集にあれば ``代表表記一致``、別名だけが載っていれば ``別名一致``。
    どちらでも無ければ用語集由来ではない（UI からの登録・手書きなど）。
    """
    rows: list[dict] = []
    for e in load_entries(dict_path):
        hit = terms.get(normalize(e["canonical"]))
        how = "代表表記一致" if hit else ""
        if hit is None:
            for alias in e["aliases"]:
                hit = terms.get(normalize(alias))
                if hit:
                    how = f"別名一致（{alias}）"
                    break
        rows.append(
            {
                "代表表記": e["canonical"],
                "カテゴリ": e["category"],
                "由来": "用語集" if hit else "その他",
                "一致": how,
                "用語集の表記": hit[0] if hit else "",
                "用語集の位置": hit[1] if hit else "",
                "別名": ", ".join(e["aliases"]),
                "部分一致": e["partial"],
                "備考": e["note"],
            }
        )
    return rows


def render(rows: list[dict], terms: dict[str, tuple[str, str]], dict_path: Path) -> str:
    """人が読む一覧テキスト。用語集由来を ★ で示す。"""
    from_glossary = [r for r in rows if r["由来"] == "用語集"]
    matched_keys = {normalize(r["用語集の表記"]) for r in from_glossary}
    unregistered = [v for k, v in terms.items() if k not in matched_keys]

    out = [
        f"辞書: {dict_path}",
        f"辞書 {len(rows)} 件のうち 用語集由来 {len(from_glossary)} 件 / "
        f"その他 {len(rows) - len(from_glossary)} 件",
        f"用語集 {len(terms)} 語のうち 辞書に未登録 {len(unregistered)} 語",
        "",
    ]
    for r in rows:
        mark = "★" if r["由来"] == "用語集" else "  "
        where = f"  ({r['用語集の位置']})" if r["用語集の位置"] else ""
        detail = f"  {r['一致']}{where}" if r["一致"] else ""
        note = f"  備考: {r['備考']}" if r["備考"] else ""
        out.append(f"{mark} {r['代表表記']}  [{r['カテゴリ']}]{detail}{note}")
    out.append("")
    out.append("★ = 用語集に載っている語（用語集由来とみなせる）")
    return "\n".join(out) + "\n"


def to_csv(rows: list[dict]) -> str:
    """CSV 文字列にする（Excel で開いてレビューする用）。"""
    if not rows:
        return ""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def _emit(text: str, out: str | None, csv_mode: bool, count: int | None = None) -> None:
    """結果をファイルか標準出力へ出す。CSV は Excel が文字化けしないよう BOM 付きで書く。"""
    if out:
        Path(out).write_text(text, encoding="utf-8-sig" if csv_mode else "utf-8")
        suffix = f"（{count} 件）" if count is not None else ""
        print(f"書き出しました: {out}{suffix}")
    else:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[union-attr]
        sys.stdout.write(text)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="用語集とマスク辞書を突き合わせ、各エントリが用語集由来かを一覧にする"
    )
    parser.add_argument(
        "glossary", help="用語集（.xlsx か、build_mask_dict の出力 .yaml）"
    )
    parser.add_argument(
        "--dict",
        default="data/mask_dict.yaml",
        help="突き合わせ先のマスク辞書（既定 data/mask_dict.yaml）",
    )
    parser.add_argument(
        "--only",
        choices=("glossary", "other"),
        help="用語集由来だけ／それ以外だけに絞る",
    )
    parser.add_argument("--csv", action="store_true", help="CSV で出力する")
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="用語集 Excel の中身を列記号つきで表示する（辞書とは突き合わせない）。"
        "「用語集 0 語」になったとき、どの列を --term-col に指定すべきか調べる用",
    )
    parser.add_argument(
        "--inspect-rows",
        type=int,
        default=12,
        help="--inspect で表示する行数（既定 12）",
    )
    parser.add_argument(
        "--out",
        help="出力先ファイル（UTF-8）。省略すると標準出力"
        "（Windows のコンソールは cp932 で化けるのでファイル推奨）",
    )
    # Excel の読み取り指定（build_mask_dict.py と同じ既定）。
    parser.add_argument("--sheet", default=None, help="シート名（既定は先頭シート）")
    parser.add_argument("--header-row", type=int, default=1, help="ヘッダ行（既定 1）")
    parser.add_argument(
        "--term-col",
        default="A",
        help="用語の列（既定 A）。カンマ区切りで複数指定できる"
        "（例 C,D,E＝用語・英語名称・日本語名称をすべて突き合わせる）",
    )
    # 種類/部分一致/大小区別の列は build_mask_dict と引数を揃えるためにあるが、
    # 突き合わせでは使わない（見るのは用語の表記だけ）。
    parser.add_argument("--kind-col", default="D", help="種類の列（既定 D。未使用）")
    parser.add_argument(
        "--partial-col", default="E", help="部分一致の列（既定 E。未使用）"
    )
    parser.add_argument(
        "--case-col", default="F", help="大小区別の列（既定 F。未使用）"
    )
    args = parser.parse_args()

    if args.inspect:
        path = Path(args.glossary)
        if not path.exists():
            print(f"用語集が見つかりません: {path}", file=sys.stderr)
            return 1
        text = inspect_glossary(
            path, args.sheet, args.header_row, args.inspect_rows, cols=10
        )
        _emit(text, args.out, csv_mode=False)
        return 0

    dict_path = Path(args.dict)
    if not dict_path.exists():
        print(f"マスク辞書が見つかりません: {dict_path}", file=sys.stderr)
        return 1

    terms = glossary_terms(args)
    if not terms:
        print(
            "用語集から 1 語も読めませんでした。--inspect で中身を確認し、"
            "--term-col / --header-row / --sheet を指定してください。",
            file=sys.stderr,
        )
    rows = classify(dict_path, terms)
    shown = rows
    if args.only == "glossary":
        shown = [r for r in rows if r["由来"] == "用語集"]
    elif args.only == "other":
        shown = [r for r in rows if r["由来"] == "その他"]

    text = to_csv(shown) if args.csv else render(shown, terms, dict_path)
    _emit(text, args.out, csv_mode=args.csv, count=len(shown))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
